'''

'''
from __future__ import print_function
from __future__ import division
from builtins import str
from builtins import object
from past.utils import old_div
from openpyxl import load_workbook
import numpy as np
import pylab
from scipy.stats import gaussian_kde
import collections
from csv import DictWriter
from collections import OrderedDict
import os.path
from scipy import stats as stats
from enum import Enum
import warnings
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles import colors

fontSize = 16


def FormatSpeciesName(species):
    ''' Formats the species name into abbreviated dot notation

    Parameters
    ----------
    species
        the species name e.g. 'Portulacaria afra'

    Returns
    -------
        the abbreviated species name e.g. 'P.afra'        
    '''
    species = str(species).strip()
    comps = species.split(' ', 1)
    if len(comps) > 1:
        abbrev_species = str.format(str("{0}.{1}"), comps[0][0], comps[1].strip())
    else:
        abbrev_species = species
    return abbrev_species

class WoodyBiomassCorrectionMethod(Enum):
    ''' Aboveground biomass correction method
    '''
    Duan = 1
    MinimumBias = 2
    NicklessZou = 3

class WoodyAbcModel(object):
    def __init__(self, model_dict={}, surrogate_dict={}, correction_method=WoodyBiomassCorrectionMethod.NicklessZou):
        ''' Object for estimating woody / above ground carbon (ABC) from allometric measurements

        Parameters
        ----------
        model_dict
            dictionary of woody allometric models
            (see table 3 in https://www.researchgate.net/publication/335531470_Aboveground_biomass_and_carbon_pool_estimates_of_Portulacaria_afra_spekboom-rich_subtropical_thicket_with_species-specific_allometric_models)
        surrogate_dict
            map from actual species to surrogate species for which models and wet/dry ratios exist
            (see supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        correction_method
            Biomass correction method to use (default WoodyBiomassCorrectionMethod.NicklessZou)
        '''
        self.model_dict = model_dict
        self.surrogate_dict = surrogate_dict
        self.correction_method = correction_method

    def Estimate(self, meas_dict={'species': '', 'canopy_length': 0., 'canopy_width': 0., 'height': 0.}):
        ''' Apply allometric model to measurements

        Parameters
        ----------
        meas_dict
            dictionary of plant species and measurements

        Returns
        -------
            ABC results in a dict (see fields in supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
                abc_dict = {'yc':0.,    # Estimated biomass (kg)
                       'yc_lc':0.,      # yc lower confidence interval (kg)
                       'yc_uc':0.,      # yc upper confidence interval (kg)
                       'height':0.,     # plant height (cm)
                       'area':0.,       # canopy area (m**2)
                       'vol':0.}        # cylindrical plant volume (m**3)
        '''
        abc_dict = {'yc':0., 'yc_lc':0., 'yc_uc':0., 'height':0., 'area':0., 'vol':0.}  # default return valued
        species =str(meas_dict['species']).strip()

        x = 0.
        CD = np.mean([meas_dict['canopy_length'], meas_dict['canopy_width']])   # canopy diameter
        CA = np.pi * (CD/2)**2                                                  # canopy area
        # Leave height in cm as other code assumes cm, but convert vol and area to m
        abc_dict['height'] = meas_dict['height']                 # pass through plant height (cm)
        abc_dict['area'] = CA/1.e4                               # canopy area (m**2)
        abc_dict['vol'] = CA*meas_dict['height']/1.e6            # cylindrical plant volume in (m**3)

        # check in surrogate_dict , then return area / vol only (yc = 0)
        if species not in self.surrogate_dict:
            warnings.warn('{0} has no key in species map, setting yc = 0'.format(species))
            return abc_dict
        allom_species = self.surrogate_dict[species]['allom_species']
        if allom_species == 'none':
            warnings.warn('{0} has no model, setting yc = 0'.format(species))
            return abc_dict
        elif allom_species not in self.model_dict:  # should never happen
            raise Exception('{0} has no key in model_dict'.format(allom_species))

        model = self.model_dict[allom_species]

        # apply allometric model
        if model['vars'] == 'CA.H':
            x = CA*meas_dict['height']
        elif model['vars'] == 'CA.SL':
            x = CA*meas_dict['height']
        elif model['vars'] == 'CD':
            x = CD
        elif model['vars'] == 'CD.H':
            x = CD*meas_dict['height']
        elif model['vars'] == 'Hgt':
            x = meas_dict['height']
        else:
            raise Exception('{0}: unknown variable', model['vars'])

        yn = np.exp(model['ay'])*x**model['by']   # "naive"

        # correct to yc
        if self.correction_method == WoodyBiomassCorrectionMethod.Duan:
            yc = yn * model['Duan']
        elif self.correction_method == WoodyBiomassCorrectionMethod.MinimumBias:
            yc = yn * model['MB']
        else:
            if yn == 0.:
                yc = 0.
            else:
                yc = np.exp(np.log(yn) + (model['sigma']**2)/2.)

        if model['use_wd_ratio']:
            wd_species = self.surrogate_dict[species]['wd_species']
            if wd_species not in self.wd_ratios:
                print('WARNING: Evalmeas_dictCs - {0} has no key in wd_ratio_dict, using 1'.format(wd_species))
            else:
                yc = yc * self.wd_ratios[wd_species]['wd_ratio']

        abc_dict['yc'] = yc * 0.48              # conversion factor from dry weight to carbon
        abc_dict['yc_lc'] = yc * model['LC']
        abc_dict['yc_uc'] = yc * model['UC']
        return abc_dict

class AgcAllometry(object):
    def __init__(self, model_file_name='', correction_method = WoodyBiomassCorrectionMethod.NicklessZou):
        ''' Class to construct allometric models and estimate AGC from woody and litter measurements

        Parameters
        ----------
        model_file_name
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
            (constructed from supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        correction_method
            Biomass correction method to use (default WoodyBiomassCorrectionMethod.NicklessZou)
        '''
        self.correction_method = correction_method
        self.__ConstructModels(model_file_name)
        self.__ConstructSurrogateMap(model_file_name)
        self.plots = {}
        self.litter_dict = {}

    def __ConstructModels(self, model_file_name):
        ''' Read in allometric models, surrogate tables and wet/dry ratios from excel file

        Parameters
        ----------
        model_file_name
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
            (constructed from supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        '''
        self.model_dict = {}
        self.master_surrogate_dict = {}

        self.model_file_name = model_file_name
        if not os.path.exists(model_file_name):
            raise Exception("Model file {0} does not exist".format(model_file_name))
        with load_workbook(self.model_file_name) as wb:

            # read in a dictionary of allometric models
            ws = wb["Allometric Models"]
            # header = [c.value for c in ws[1]]
            for r in ws[2:ws.max_row]:
                if r[0].value is None:
                    break
                species = FormatSpeciesName('{0} {1}'.format(str(r[0].value).strip(), str(r[1].value).strip()))
                model = {'vars': r[3].value, 'ay': r[6].value, 'by': r[7].value, 'sigma': r[8].value, 'LC': r[9].value,
                         'UC': r[10].value, 'Duan': r[12].value, 'MB': r[13].value,
                         'use_wd_ratio': False if (str(r[14].value) == 'x') else True}
                self.model_dict[species] = model
            ws = None

            # read in a dictionary of wet/dry ratios
            ws = wb["Wet Dry Ratios"]
            self.wd_ratio_dict = {}
            for r in ws[2:ws.max_row]:
                if r[0].value is None:
                    break
                species = FormatSpeciesName('{0} {1}'.format(str(r[0].value).strip(), str(r[1].value).strip()))
                model = {'wd_species': species, 'wd_ratio': r[4].value}
                self.wd_ratio_dict[species] = model


    def __ConstructSurrogateMap(self, model_file_name):
        ''' Construct master surrogate species map from combination of Cos Bolus and Marius van der Vyver contributions

        Parameters
        ----------
        model_file_name
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
        '''
        import copy
        self.model_file_name = model_file_name
        if not os.path.exists(model_file_name):
            raise Exception("Model file {0} does not exist".format(model_file_name))

        with load_workbook(self.model_file_name) as wb:
            # read in a Marius van der Vyver's surrogate map
            ws = wb["Surrogates"]
            # header = [c.value for c in ws[1]]
            self.mvdv_surrogate_dict = {}
            for r in ws[2:ws.max_row]:
                if r[0].value is None:
                    break

                species = FormatSpeciesName(str(r[0].value).strip())
                model = {'species': species, 'allom_species': FormatSpeciesName(str(r[1].value).strip()),
                         FormatSpeciesName(str(r[2].value).strip()) if r[2].value is not None else None}
                self.mvdv_surrogate_dict[species] = model

            self.cb_surrogate_dict = {}

            # read in a Cos Bolus' surrogate map
            ws = wb.get_sheet_by_name("CB Surrogates")
            for r in ws[2:ws.max_row]:
                species = str(r[1].value).strip()

                for c in r[6:]:     # check spreadsheet format
                    if (c.value is not None) and (c.value != ''):
                        map_species = str(c.value).strip()
                        break

                # add wet/dry surrogate
                wd_species = None
                if species in self.wd_ratio_dict:
                    wd_species = species
                elif model['allom_species'] in self.wd_ratio_dict:
                    wd_species = model['allom_species']

                model = {'species': species, 'allom_species': FormatSpeciesName(map_species), 'wd_species': wd_species,
                        'species_full': str(r[0].value).strip()}  # hack for translating between name formats
                self.cb_surrogate_dict[species] = model

        # combine the two species maps into one
        self.master_surrogate_dict = {}
        # copy across Marius' map
        for species, surrogate_species_dict in self.mvdv_surrogate_dict.items():
            # use abbrev name as the key always
            self.master_surrogate_dict[species] = copy.deepcopy(surrogate_species_dict)
            # do some error checking
            if surrogate_species_dict['allom_species'] not in self.model_dict:
                warnings.warn('Species: {0}, allom surrogate: {1} - No allometric model'.format(species, surrogate_species_dict['allom_species']))
            else:
                if self.model_dict[surrogate_species_dict['allom_species']]['use_wd_ratio']:
                    if surrogate_species_dict['wd_species'] is not None:
                        if surrogate_species_dict['wd_species'] not in self.wd_ratio_dict:
                            warnings.warn('Species: {0}, wd surrogate: {1} - No wet:dry ratio'.format(species,
                                                                                                      surrogate_species_dict['wd_species']))
                            self.master_surrogate_dict[species]['wd_species'] = None
                    else:
                        warnings.warn('Species: {0}, no wd surrogate'.format(species))

        # now add Cos' map, with wet/dry ratios
        for species, surrogate_species_dict in self.cb_surrogate_dict.items():
            if species not in self.master_surrogate_dict:
                self.master_surrogate_dict[species] = copy.deepcopy(surrogate_species_dict)

                if surrogate_species_dict['allom_species'] == 'none':  # this has no surrogate at all and is excluded
                    self.master_surrogate_dict[species]['wd_species'] = None
                elif species in self.wd_ratio_dict: # if the source species has wd_ratio then refer to this directly
                    self.master_surrogate_dict[species]['wd_species'] = species
                elif surrogate_species_dict['allom_species'] in self.wd_ratio_dict: # else if the allom species has wd_ratio then refer to this
                    self.master_surrogate_dict[species]['wd_species'] = surrogate_species_dict['allom_species']

                elif self.model_dict[surrogate_species_dict['allom_species']]['use_wd_ratio']: # else check we actually need a wd ratio
                    # look up wd_species in surrogate map for allom surrogate if possible
                    if surrogate_species_dict['allom_species'] in self.mvdv_surrogate_dict and self.mvdv_surrogate_dict[surrogate_species_dict['allom_species']]['wd_species'] is not None:
                        self.master_surrogate_dict[species]['wd_species'] = self.mvdv_surrogate_dict[surrogate_species_dict['allom_species']]['wd_species']
                    else:
                        warnings.warn('Species: {0}, allom surrogate: {1} - No wd surrogate'.format(species, surrogate_species_dict['allom_species']))

    def __ReadLitter(self, litter_file_name=None):
        ''' Read excel file into dict of dry litter weight for each plot

        Parameters
        ----------
        litter_file_name
            name of excel file containing plot ID's and litter weights
        '''

        if litter_file_name is None:
            litter_file_name = self.woody_file_name
        self.litter_dict = {}
        if not os.path.exists(litter_file_name):
            raise Exception("Litter file {0} does not exist".format(litter_file_name))

        with load_workbook(litter_file_name, data_only=True) as wb:
            ws = wb['Final Litter_copied']
            for r in ws[2:ws.max_row]:      # loop through plots/rows
                plot_id = str(r[0].value).strip().upper()   # parse plot ID
                plot_id = plot_id.replace('-0', '')
                plot_id = plot_id.replace('-', '')
                if plot_id == '' or plot_id is None or plot_id == 'NONE' or r[1].value == 0:
                    warnings.warn('Empty record, continuing...')
                    continue
                if not np.isreal(r[1].value) or r[1].value is None:
                    dry_weight = 0.
                    warnings.warn('No data for plot {0}'.format(plot_id))        # these are typically excluded / not sampled plots
                else:
                    dry_weight = r[1].value

                if plot_id in self.litter_dict:
                    self.litter_dict[plot_id]['dry_weight'] += dry_weight
                    warnings.warn('Multiple values for plot {0}'.format(plot_id))
                else:
                    self.litter_dict[plot_id] = {'dry_weight': dry_weight}

    def EstimatePlantAbc(self, woody_file_name='', make_marked_file=False):
        ''' Estimate aboveground biomass carbon (ABC) for each plant in each plot

        Parameters
        ----------
        woody_file_name
            excel file containing plant measurements for each plot
        make_marked_file
            create an output excel file that highlighs problematic rows in woody_file_name
        '''

        woody_abc_model = WoodyAbcModel(model_dict=self.model_dict, surrogate_dict=self.master_surrogate_dict,
                                        correction_method=self.correction_method)
        if make_marked_file:
            ok_colour = Color(auto=True)

        with load_workbook(woody_file_name) as wb:
            ws = wb.get_sheet_by_name("Consolidated data")

            self.unmodelled_species = {'unknown': {}, 'none': {}}       # keep a record of species without models
            self.plots = collections.OrderedDict()
            for r in ws[2:ws.max_row]:
                if r is None or r[2].value is None:
                    continue
                species = str(r[3].value).strip()

                plot_id = str(r[0].value).strip()
                dashLoc = str(plot_id).find('-')
                if dashLoc < 0:
                    dashLoc = 2
                plot_id = plot_id.replace('-', '').upper()
                id_num = np.int32(plot_id[dashLoc:])  # get rid of leading zeros
                plot_id = '%s%d' % (plot_id[:dashLoc], id_num)
                plot_size = np.int32(str(r[1].value).lower().split('x')[0])
                meas_dict = OrderedDict({'ID': plot_id, 'degr_class': str(r[2].value), 'orig_species': species,
                                      'canopy_width': r[4].value, 'canopy_length': r[5].value, 'height': r[6].value,
                                      'species': species, 'plot_size': plot_size})

                # error checking
                fields = ['canopy_width', 'canopy_length', 'height']
                fields_ok = True
                for f in fields:
                    if meas_dict[f] is None:
                        warnings.warn('ID: {0}, species: {1}, has incomplete data'.format(plot_id, species))
                        meas_dict[f] = 0
                        fields_ok = False

                meas_dict['bsd'] = str(r[7].value) if (r.__len__() > 7) else ''
                abc_dict = woody_abc_model.Estimate(meas_dict)

                for key in list(abc_dict.keys()):   # copy to meas_dict
                    meas_dict[key] = abc_dict[key]

                if make_marked_file:    # mark problem cells
                    if species not in self.master_surrogate_dict or not fields_ok:
                        r[3].fill = PatternFill(fgColor=colors.YELLOW, fill_type='solid')
                        print('Marking row {0}'.format(r[3].row))
                    else:
                        r[3].fill = PatternFill(fgColor=ok_colour, fill_type='solid',)

                # error checking done in woody_abc_model.Estimate, but this remembers species without models
                if species not in self.master_surrogate_dict or self.master_surrogate_dict[species]['allom_species'] == 'none':
                    if species not in self.master_surrogate_dict:
                        key = 'unknown'
                    elif self.master_surrogate_dict[species]['allom_species'] == 'none':
                        key = 'none'
                    if species in self.unmodelled_species:
                        self.unmodelled_species[key][species]['count'] += 1
                        self.unmodelled_species[key][species]['vol'] += old_div(abc_dict['vol'],1.e6)
                    else:
                        self.unmodelled_species[key][species] = {}
                        self.unmodelled_species[key][species]['count'] = 1
                        self.unmodelled_species[key][species]['vol'] = old_div(abc_dict['vol'], 1.e6)

                if plot_id in self.plots:
                    self.plots[plot_id].append(meas_dict)
                else:
                    self.plots[plot_id] = [meas_dict]

            if make_marked_file:
                out_file_name = str.format('{0}/{1}_Marked.xlsx', os.path.dirname(woody_file_name),
                                           os.path.splitext(os.path.basename(woody_file_name))[0])
                wb.save(filename=out_file_name)

        print('Unknown species:')
        for k, v in self.unmodelled_species['unknown'].items():
            print(k, v)
        print('Unmodelled species:')
        for k, v in self.unmodelled_species['none'].items():
            print(k, v)

    # @staticmethod
    def EvalPlotSummaryCs(self):
        if len(self.plots) == 0:
            print('EstimatePlantAbc has not been called')
            return
        self.summary_plots = {}
        # summing yc_lc and yc_uc is probably not valid - summing would average out errors and reduce lc and uc
        # fields_to_summarise = ['yc', 'yc_lc', 'yc_uc', 'height', 'vol']
        fields_to_summarise = ['yc', 'height', 'vol']
        for id, plot in self.plots.items():
            plot_sizes = np.array([record['plot_size'] for record in plot])
            plot_sizes_un = np.unique(plot_sizes)
            vectors_to_summarise = {}
            vector_dict = {}
            for field in fields_to_summarise:
                vectors_to_summarise[field] = np.array([record[field] for record in plot])
                vector_dict[field] = {'v': vectors_to_summarise[field]}

            heights = np.array([record['height'] for record in plot])
            # ycs = np.array([record['yc'] for record in plot])
            # yc_lcs = np.array([record['yc_lc'] for record in plot])
            # yc_ucs = np.array([record['yc_uc'] for record in plot])
            # vols = np.array([record['vol'] for record in plot])
            # vector_dict = {'yc': {'v': ycs}, 'yc_lc': {'v': yc_lcs},'yc_uc': {'v': yc_ucs}, 'vol': {'v': vols}, 'height': {'v': heights}}
            if plot_sizes_un.size > 1:      # it is a nested plot, so we need to extrapolate
                nest_record_idx = plot_sizes == plot_sizes_un.min()
                small_record_idx = heights < 50

                for k, v in vector_dict.items():
                    vv = v['v']
                    nest_v = vv[nest_record_idx & ~small_record_idx]
                    small_v = vv[nest_record_idx & small_record_idx]
                    out_v = vv[~nest_record_idx]
                    sum_v = out_v.sum() + nest_v.sum() + (small_v.sum() * ((old_div(plot_sizes_un.max(), plot_sizes_un.min())) ** 2))
                    mean_v = out_v.tolist() + nest_v.tolist() + (small_v.tolist() * ((old_div(plot_sizes_un.max(), plot_sizes_un.min())) ** 2))
                    v['sum_v'] = sum_v
                    v['mean_v'] = np.array(mean_v).mean()
                    v['n'] = out_v.size + nest_v.size + (small_v.size * (old_div(plot_sizes_un.max(), plot_sizes_un.min())) ** 2)
            else:
                for k, v in vector_dict.items():
                    vv = v['v']
                    v['sum_v'] = vv.sum()
                    v['mean_v'] = vv.mean()
                    v['n'] = vv.size

            summary_plot = OrderedDict()
            summary_plot['ID'] = id
            summary_plot['Stratum'] = plot[0]['degr_class']
            summary_plot['Size'] = plot_sizes_un.max()
            summary_plot['N'] = vector_dict['height']['n']
            summary_plot['Vol'] = vector_dict['vol']['sum_v']
            summary_plot['VolHa'] = old_div((100. ** 2) * vector_dict['vol']['sum_v'], (plot_sizes_un.max() ** 2))
            summary_plot['Height'] = vector_dict['height']['sum_v']
            summary_plot['HeightHa'] = old_div((100. ** 2) * vector_dict['height']['sum_v'], (plot_sizes_un.max() ** 2))
            summary_plot['MeanHeight'] = vector_dict['height']['mean_v']
            summary_plot['Abc'] = vector_dict['yc']['sum_v']
            summary_plot['AbcHa'] = old_div((100. ** 2) * vector_dict['yc']['sum_v'], (plot_sizes_un.max() ** 2))
            # summary_plot['AbcLc'] = vector_dict['yc_lc']['sum_v']
            # summary_plot['AbcLcHa'] = (100. ** 2) * vector_dict['yc_lc']['sum_v'] / (plot_sizes_un.max() ** 2)
            # summary_plot['AbcUc'] = vector_dict['yc_uc']['sum_v']
            # summary_plot['AbcUcHa'] = (100. ** 2) * vector_dict['yc_uc']['sum_v'] / (plot_sizes_un.max() ** 2)

            if len(self.litter_dict) == 0:
                print('WARNING: EvalPlotSummaryCs - no litter data')
            else:
                if id in self.litter_dict and self.litter_dict[id]['dry_weight'] > 0.:
                    summary_plot['LitterC'] = old_div(0.48 * self.litter_dict[id]['dry_weight'], 1000)  # g to kg
                    # the litter quadrats are uniform across all plot sizes.  the dry weight is converted to carbon using the factor 0.48 (according to Cos)
                    # or 0.37 (according to James / CDM)
                    # current the trial plots have no litter data (INT* and TCH*)
                    # NB note, there is a factor of 44/12 in the CDM equations that I don't quite understand.  I think it is a conversion from C weight to
                    # CO2 weight (i.e. a chemical thing that represents not how much carbon is stored by how much C02 was captured out the atmosphere).
                    # we should make sure that we are using the same units in the woody and litter C i.e. we need to check what units Marius eq give
                    # Mills and van der Vyver use .48
                    summary_plot['LitterCHa'] = old_div(summary_plot['LitterC'] * (100. ** 2), (4 * (0.5 ** 2)))
                    # summary_plot['LitterCHa'] = summary_plot['LitterC'] * 0.37 * (100. ** 2) / (4 * (0.5 ** 2))
                    summary_plot['AgcHa'] = summary_plot['LitterCHa'] + summary_plot['AbcHa']
                else:
                    print('WARNING: EvalPlotSummaryCs - no litter data for ID: {0}'.format(id))
                    summary_plot['LitterC'] = np.nan
                    summary_plot['LitterCHa'] = np.nan
                    summary_plot['AgcHa'] = summary_plot['AbcHa']

            self.summary_plots[id] = summary_plot

    def WriteAllCsFile(self, out_file_name=None):
        if len(self.plots) == 0:
            print('EstimatePlantAbc has not been called')
            return

        if out_file_name is None:
            out_file_name = str.format(str('{0}/{1} - All WoodyC.csv'), os.path.dirname(self.woody_file_name),
                                     os.path.splitext(os.path.basename(self.woody_file_name))[0])
        print('Writing plot data to: {0}'.format(out_file_name))
        with open(out_file_name,'w', newline='') as outfile:
            writer = DictWriter(outfile, list(list(self.plots.values())[0][0].keys()))
            writer.writeheader()
            for plot in list(self.plots.values()):
                writer.writerows(plot)

    def WriteSummaryFile(self, out_file_name=None):
        if len(self.summary_plots) == 0:
            print('EvalPlotSummaryCs has not been called')
            return

        if out_file_name is None:
            out_file_name = str.format(str('{0}/{1} - Summary WoodyC & LitterC.csv'), os.path.dirname(self.woody_file_name),
                                     os.path.splitext(os.path.basename(self.woody_file_name))[0])
        print('Writing plot summary to: {0}'.format(out_file_name))
        with open(out_file_name, 'w', newline='') as outfile:
            writer = DictWriter(outfile, list(self.summary_plots.values())[0].keys())
            writer.writeheader()
            writer.writerows(list(self.summary_plots.values()))
