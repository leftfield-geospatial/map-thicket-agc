"""
    GEF5-SLM: Above ground carbon estimation in thicket using multi-spectral images
    Copyright (C) 2020 Dugal Harris
    Email: dugalh@gmail.com

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU Affero General Public License as
    published by the Free Software Foundation, either version 3 of the
    License, or any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU Affero General Public License for more details.

    You should have received a copy of the GNU Affero General Public License
    along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""


from builtins import str
from openpyxl import load_workbook
import numpy as np
from collections import OrderedDict
import os.path
from enum import Enum
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles import colors
import pandas as pd
from map_thicket_agc import get_logger

logger = get_logger(__name__)

# globals
biomass_to_carbon_w = 0.48      # factor to convert from biomass to carbon weight
# (Note that CDM and J Reeler use 0.37, while C Bolus, M vd Vyver and A Mills use 0.48)
nested_height_thresh = 50.      # cutoff plant height for nested plot (cm)

def format_species_name(species):
    """ Formats the species name into abbreviated dot notation.

    Parameters
    ----------
    species : str
        the species name e.g. 'Portulacaria afra'

    Returns
    -------
    str
        the abbreviated species name e.g. 'P.afra'
    """
    species = str(species).strip()
    comps = species.split(' ', 1)
    if len(comps) > 1:
        abbrev_species = str.format(str("{0}.{1}"), comps[0][0], comps[1].strip())
    else:
        abbrev_species = species
    return abbrev_species

class BiomassCorrectionMethod(Enum):
    """ Woody biomass correction method.
    """
    Duan = 1
    MinimumBias = 2
    NicklessZou = 3

class AbcPlantEstimator:
    def __init__(self, model_dict=None, surrogate_dict=None, wd_ratio_dict=None, correction_method=BiomassCorrectionMethod.NicklessZou):
        """ Class for estimating woody / above ground carbon (ABC) from allometric measurements.

        Parameters
        ----------
        model_dict: dict
            woody allometric models
            (see table 3 in https://www.researchgate.net/publication/335531470_Aboveground_biomass_and_carbon_pool_estimates_of_Portulacaria_afra_spekboom-rich_subtropical_thicket_with_species-specific_allometric_models)
        surrogate_dict: dict
            map from actual species to surrogate species for which models and wet/dry ratios exist
            (see supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        correction_method: BiomassCorrectionMethod
            Biomass correction method to use (default BiomassCorrectionMethod.NicklessZou)
        """
        self.model_dict = model_dict
        self.surrogate_dict = surrogate_dict
        self.wd_ratio_dict = wd_ratio_dict
        self.correction_method = correction_method

    def estimate(self, meas_dict=None):
        """ Apply allometric model to measurements

        Parameters
        ----------
        meas_dict : dict
            plant species and measurements in form {'species': '', 'canopy_length': 0., 'canopy_width': 0., 'height': 0.}

        Returns
        -------
            ABC results in a dict (see fields in supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
                abc_dict = {'yc':0.,    # Estimated biomass (kg)
                       'yc_lc':0.,      # yc lower confidence interval (kg)
                       'yc_uc':0.,      # yc upper confidence interval (kg)
                       'height':0.,     # plant height (cm)
                       'area':0.,       # canopy area (m**2)
                       'vol':0.}        # cylindrical plant volume (m**3)
        """
        abc_dict = {'yc':0., 'yc_lc':0., 'yc_uc':0., 'height':0., 'area':0., 'vol':0.}  # default return valued
        species =str(meas_dict['species']).strip()

        CD = np.mean([meas_dict['canopy_length'], meas_dict['canopy_width']])   # canopy diameter
        CA = np.pi * (CD/2)**2                                                  # canopy area
        # Leave height in cm as other code assumes cm, but convert vol and area to m
        abc_dict['height'] = meas_dict['height']                 # pass through plant height (cm)
        abc_dict['area'] = CA/1.e4                               # canopy area (m**2)
        abc_dict['vol'] = CA*meas_dict['height']/1.e6            # cylindrical plant volume in (m**3)

        # check in surrogate_dict , then return area / vol only (yc = 0)
        if species not in self.surrogate_dict:
            logger.warning('{0} has no key in species map, setting yc = 0'.format(species))
            return abc_dict
        allom_species = self.surrogate_dict[species]['allom_species']
        if allom_species == 'none':
            logger.warning('{0} has no model, setting yc = 0'.format(species))
            return abc_dict
        elif allom_species not in self.model_dict:  # should never happen
            raise Exception('{0} has no key in model_dict'.format(allom_species))

        model = self.model_dict[allom_species]

        # apply allometric model
        if model['vars'] == 'CA.H':
            x = CA*meas_dict['height']
        elif model['vars'] == 'CA.SL':
            x = CA*meas_dict['height']
        elif model['vars'] == 'CD':
            x = CD
        elif model['vars'] == 'CD.H':
            x = CD*meas_dict['height']
        elif model['vars'] == 'Hgt':
            x = meas_dict['height']
        else:
            raise Exception('{0}: unknown variable', model['vars'])

        yn = np.exp(model['ay'])*x**model['by']   # "naive"

        # correct to yc
        if self.correction_method == BiomassCorrectionMethod.Duan:
            yc = yn * model['Duan']
        elif self.correction_method == BiomassCorrectionMethod.MinimumBias:
            yc = yn * model['MB']
        else:
            if yn == 0.:
                yc = 0.
            else:
                yc = np.exp(np.log(yn) + (model['sigma']**2)/2.)

        if model['use_wd_ratio']:
            wd_species = self.surrogate_dict[species]['wd_species']
            if wd_species not in self.wd_ratio_dict:
                logger.warning('{0} has no key in wd_ratio_dict, using 1'.format(wd_species))
            else:
                yc = yc * self.wd_ratio_dict[wd_species]['wd_ratio']

        abc_dict['yc'] = yc * biomass_to_carbon_w              # conversion factor from dry weight to carbon
        abc_dict['yc_lc'] = yc * model['LC']
        abc_dict['yc_uc'] = yc * model['UC']
        return abc_dict

class AbcAggregator:
    def __init__(self, model_file_name='', correction_method = BiomassCorrectionMethod.NicklessZou):
        """ Class to co-ordinate estimation of ABC for all plant measurements in a file.

        Parameters
        ----------
        model_file_name : str
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
            (constructed from supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        correction_method : BiomassCorrectionMethod
            Biomass correction method to use (default WoodyBiomassCorrectionMethod.NicklessZou)
        """
        self.correction_method = correction_method
        self._construct_models(model_file_name)
        self._construct_surrogate_map(model_file_name)

        self.plot_abc_df = pd.DataFrame()
        self._unmodelled_species = {}
        self._woody_file_name = None

    def _construct_models(self, model_file_name):
        """ Read in allometric models, surrogate tables and wet/dry ratios from excel file

        Parameters
        ----------
        model_file_name : str
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
            (constructed from supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        """
        self.model_dict = {}
        self.master_surrogate_dict = {}

        self.model_file_name = model_file_name
        if not os.path.exists(model_file_name):
            raise Exception("Model file {0} does not exist".format(model_file_name))
        wb = load_workbook(self.model_file_name)
        try:
            # read in a dictionary of allometric models
            ws = wb["Allometric Models"]
            # header = [c.value for c in ws[1]]
            for r in ws[2:ws.max_row]:
                if r[0].value is None:
                    break
                species = format_species_name('{0} {1}'.format(str(r[0].value).strip(), str(r[1].value).strip()))
                model = {'vars': r[3].value, 'ay': r[6].value, 'by': r[7].value, 'sigma': r[8].value, 'LC': r[9].value,
                         'UC': r[10].value, 'Duan': r[12].value, 'MB': r[13].value,
                         'use_wd_ratio': False if (str(r[14].value) == 'x') else True}
                self.model_dict[species] = model

            # read in a dictionary of wet/dry ratios
            ws = wb["Wet Dry Ratios"]
            self.wd_ratio_dict = {}
            for r in ws[2:ws.max_row]:
                if r[0].value is None:
                    break
                species = format_species_name('{0} {1}'.format(str(r[0].value).strip(), str(r[1].value).strip()))
                model = {'wd_species': species, 'wd_ratio': r[4].value}
                self.wd_ratio_dict[species] = model
        finally:
            wb.close()

    def _construct_surrogate_map(self, model_file_name):
        """ Construct master surrogate species map from combination of Cos Bolus and Marius van der Vyver contributions

        Parameters
        ----------
        model_file_name : str
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
        """
        import copy
        self.model_file_name = model_file_name
        if not os.path.exists(model_file_name):
            raise Exception("Model file {0} does not exist".format(model_file_name))

        wb = load_workbook(self.model_file_name)
        try:
            # read in a Marius van der Vyver's surrogate map
            ws = wb["Surrogates"]
            # header = [c.value for c in ws[1]]
            self.mvdv_surrogate_dict = {}
            for r in ws[2:ws.max_row]:
                if r[0].value is None:
                    break

                species = format_species_name(str(r[0].value).strip())
                model = {'species': species, 'allom_species': format_species_name(str(r[1].value).strip()),
                         'wd_species': format_species_name(str(r[2].value).strip()) if r[2].value is not None else None}
                self.mvdv_surrogate_dict[species] = model

            self.cb_surrogate_dict = {}

            # read in a Cos Bolus' surrogate map
            ws = wb.get_sheet_by_name("CB Surrogates")
            for r in ws[2:ws.max_row]:
                species = str(r[1].value).strip()
                map_species = ''
                for c in r[6:]:     # check spreadsheet format
                    if (c.value is not None) and (c.value != ''):
                        map_species = str(c.value).strip()
                        break
                allom_species = format_species_name(map_species)
                # add wet/dry surrogate
                wd_species = None
                if species in self.wd_ratio_dict:
                    wd_species = species
                elif allom_species in self.wd_ratio_dict:
                    wd_species = allom_species

                model = {'species': species, 'allom_species': allom_species, 'wd_species': wd_species,
                        'species_full': str(r[0].value).strip()}  # for translating between name formats
                self.cb_surrogate_dict[species] = model
        finally:
            wb.close()

        # combine the two species maps into one
        self.master_surrogate_dict = {}
        # copy across Marius' map
        for species, surrogate_species_dict in self.mvdv_surrogate_dict.items():
            # use abbrev name as the key always
            self.master_surrogate_dict[species] = copy.deepcopy(surrogate_species_dict)
            # do some error checking
            if surrogate_species_dict['allom_species'] not in self.model_dict:
                logger.warning('Species: {0}, allom surrogate: {1} - No allometric model'.format(species, surrogate_species_dict['allom_species']))
            else:
                if self.model_dict[surrogate_species_dict['allom_species']]['use_wd_ratio']:
                    if surrogate_species_dict['wd_species'] is not None:
                        if surrogate_species_dict['wd_species'] not in self.wd_ratio_dict:
                            logger.warning('Species: {0}, wd surrogate: {1} - No wet:dry ratio'.format(species,
                                                                                                      surrogate_species_dict['wd_species']))
                            self.master_surrogate_dict[species]['wd_species'] = None
                    else:
                        logger.warning('Species: {0}, no wd surrogate'.format(species))

        # now add Cos' map, with wet/dry ratios
        for species, surrogate_species_dict in self.cb_surrogate_dict.items():
            if species not in self.master_surrogate_dict:
                self.master_surrogate_dict[species] = copy.deepcopy(surrogate_species_dict)

                if surrogate_species_dict['allom_species'] == 'none':  # this has no surrogate at all and is excluded
                    self.master_surrogate_dict[species]['wd_species'] = None
                elif species in self.wd_ratio_dict: # if the source species has wd_ratio then refer to this directly
                    self.master_surrogate_dict[species]['wd_species'] = species
                elif surrogate_species_dict['allom_species'] in self.wd_ratio_dict: # else if the allom species has wd_ratio then refer to this
                    self.master_surrogate_dict[species]['wd_species'] = surrogate_species_dict['allom_species']

                elif self.model_dict[surrogate_species_dict['allom_species']]['use_wd_ratio']: # else check we actually need a wd ratio
                    # look up wd_species in surrogate map for allom surrogate if possible
                    if surrogate_species_dict['allom_species'] in self.mvdv_surrogate_dict and self.mvdv_surrogate_dict[surrogate_species_dict['allom_species']]['wd_species'] is not None:
                        self.master_surrogate_dict[species]['wd_species'] = self.mvdv_surrogate_dict[surrogate_species_dict['allom_species']]['wd_species']
                    else:
                        logger.warning('Species: {0}, allom surrogate: {1} - No wd surrogate'.format(species, surrogate_species_dict['allom_species']))


    def aggregate(self, woody_file_name='', make_marked_file=False):
        """ Estimate aboveground biomass carbon (ABC) for each plant in each plot.
         Plant measurements are read from excel spreadsheet of field data.

        Parameters
        ----------
        woody_file_name : str
            excel file containing plant measurements for each plot
        make_marked_file : bool
            create an output excel file that highlights problematic rows in woody_file_name (default = False)

        Returns
        -------
            a dict of plant ABC etc values
        """
        abc_plant_estimator = AbcPlantEstimator(model_dict=self.model_dict, surrogate_dict=self.master_surrogate_dict,
                                                wd_ratio_dict=self.wd_ratio_dict, correction_method=self.correction_method)
        ok_colour = Color(auto=True)

        wb = load_workbook(woody_file_name)
        try:
            self._woody_file_name = woody_file_name
            ws = wb.get_sheet_by_name("Consolidated data")

            self._unmodelled_species = {'unknown': {}, 'none': {}}       # keep a record of species without models
            plot_abc_list = []
            for r in ws[2:ws.max_row]:      # loop through each plant
                if r is None or r[2].value is None:
                    continue
                species = str(r[3].value).strip()

                # parse plot ID
                plot_id = str(r[0].value).strip()
                dashLoc = 2 if str(plot_id).find('-') < 0 else str(plot_id).find('-')
                plot_id = plot_id.replace('-', '').upper()
                id_num = np.int32(plot_id[dashLoc:])  # get rid of leading zeros
                plot_id = '%s%d' % (plot_id[:dashLoc], id_num)

                plot_size = np.int32(str(r[1].value).lower().split('x')[0])
                degr_class = str(r[2].value).strip()
                if degr_class == 'Degraded':
                    degr_class = 'Severe'
                elif degr_class == 'Pristine':
                    degr_class = 'Intact'

                meas_dict = OrderedDict({'ID': plot_id, 'degr_class': degr_class, 'orig_species': species,
                                      'canopy_width': r[4].value, 'canopy_length': r[5].value, 'height': r[6].value,
                                      'species': species, 'plot_size': plot_size})

                # error checking
                fields = ['canopy_width', 'canopy_length', 'height']
                fields_ok = True
                for f in fields:
                    if meas_dict[f] is None:
                        logger.warning('ID: {0}, species: {1}, has incomplete data'.format(plot_id, species))
                        meas_dict[f] = 0
                        fields_ok = False

                meas_dict['bsd'] = str(r[7].value) if (r.__len__() > 7) else ''
                abc_dict = abc_plant_estimator.estimate(meas_dict)

                for key in list(abc_dict.keys()):   # copy to meas_dict
                    meas_dict[key] = abc_dict[key]

                if make_marked_file:    # mark problem cells in excel spreadsheet
                    if species not in self.master_surrogate_dict or not fields_ok:
                        r[3].fill = PatternFill(fgColor=colors.COLOR_INDEX[5], fill_type='solid')
                        logger.debug('Marking row {0}'.format(r[3].row))
                    else:
                        r[3].fill = PatternFill(fgColor=ok_colour, fill_type='solid',)

                # gather stats on species without models
                if species not in self.master_surrogate_dict or self.master_surrogate_dict[species]['allom_species'] == 'none':
                    key = ''
                    if species not in self.master_surrogate_dict:
                        key = 'unknown'     # unknown unknowns
                    elif self.master_surrogate_dict[species]['allom_species'] == 'none':
                        key = 'none'        # known unknowns
                    if species in self._unmodelled_species:
                        self._unmodelled_species[key][species]['count'] += 1
                        self._unmodelled_species[key][species]['vol'] += abc_dict['vol'] / 1.e6
                    else:
                        self._unmodelled_species[key][species] = {'count': 1, 'vol': abc_dict['vol'] / 1.e6}

                plot_abc_list.append(meas_dict)

            self.plot_abc_df = pd.DataFrame(plot_abc_list)
            if make_marked_file:
                out_file_name = str.format('{0}/{1}_Marked.xlsx', os.path.dirname(woody_file_name),
                                           os.path.splitext(os.path.basename(woody_file_name))[0])
                wb.save(filename=out_file_name)
        finally:
            wb.close()

        logger.info('Unknown species:')
        for k, v in self._unmodelled_species['unknown'].items():
            logger.info(f'{k} {v}')
        logger.info('Unmodelled species:')
        for k, v in self._unmodelled_species['none'].items():
            logger.info(f'{k} {v}')

        return self.plot_abc_df

    def write_file(self, out_file_name=None):
        if len(self.plot_abc_df) == 0:
            raise Exception("There is no ABC data - call aggregate first")

        if out_file_name is None:
            out_file_name = str.format(str('{0}/{1} - All WoodyC.csv'), os.path.dirname(self._woody_file_name),
                                       os.path.splitext(os.path.basename(self._woody_file_name))[0])

        self.plot_abc_df.to_csv(out_file_name, index=False)

class AgcPlotEstimator:
    def __init__(self, model_file_name='', correction_method=BiomassCorrectionMethod.NicklessZou):
        """ Class to estimate aboveground carbon (AGC) per plot.  Woody ABC and Litter C are aggregated.

        Parameters
        ----------
        model_file_name : str
            excel spreadsheet specifying allometric models, surrogate tables and wet/dry ratios
            (constructed from supplementary data in https://doi.org/10.1016/j.foreco.2019.05.048)
        correction_method : BiomassCorrectionMethod
            Biomass correction method to use (default BiomassCorrectionMethod.NicklessZou)
        """
        self.plot_summary_agc_df = pd.DataFrame()
        self.abc_aggregator = AbcAggregator(model_file_name=model_file_name, correction_method=correction_method)
        self._plot_litter_df = pd.DataFrame()
        self._woody_file_name = None

    def _read_litter(self, litter_file_name=None):
        """ Read excel file into dict of dry litter weight for each plot

        Parameters
        ----------
        litter_file_name : str
            name of excel file containing plot ID's and litter weights
        """

        if litter_file_name is None:
            litter_file_name = self._woody_file_name
        plot_litter_dict = {}
        if not os.path.exists(litter_file_name):
            raise Exception("Litter file {0} does not exist".format(litter_file_name))

        wb = load_workbook(litter_file_name, data_only=True)
        try:
            ws = wb['Final Litter_copied']
            for r in ws[2:ws.max_row]:      # loop through plots/rows
                # parse plot ID
                plot_id = str(r[0].value).strip().upper()
                plot_id = plot_id.replace('-0', '')
                plot_id = plot_id.replace('-', '')
                if plot_id == '' or plot_id is None or plot_id == 'NONE' or r[1].value == 0:
                    logger.warning('Empty record, continuing...')
                    continue

                if not np.isreal(r[1].value) or r[1].value is None:
                    dry_weight = 0.
                    logger.warning('No data for plot {0}'.format(plot_id))        # these are typically excluded / not sampled plots
                else:
                    dry_weight = r[1].value
                if plot_id in plot_litter_dict:
                    plot_litter_dict[plot_id]['dry_weight'] += dry_weight
                    logger.warning('Multiple values for plot {0}'.format(plot_id))
                else:
                    plot_litter_dict[plot_id] = {'dry_weight': dry_weight}
        finally:
            wb.close()
            self._plot_litter_df = pd.DataFrame.from_dict(plot_litter_dict, orient='index')
            self._plot_litter_df['ID'] = self._plot_litter_df.index

    def estimate(self, woody_file_name='', litter_file_name='', make_marked_file=False):
        """ Estimate total AGC per plot

        Parameters
        ----------
        woody_file_name : str
            excel file containing plant measurements for each plot
        litter_file_name : str
            name of excel file containing plot ID's and litter weights
        make_marked_file : bool
            create an output excel file that highlights problematic rows in woody_file_name (default = False)

        Returns
        -------
            a dict of plot AGC etc values
        """
        plot_summary_agc_dict = {}
        self.plot_summary_agc_df = pd.DataFrame()
        self.abc_aggregator.aggregate(woody_file_name=woody_file_name, make_marked_file=make_marked_file)
        self._woody_file_name = woody_file_name
        self._plot_litter_df = pd.DataFrame()
        self._read_litter(litter_file_name)

        fields_to_summarise = ['yc', 'height', 'vol']
        for plot_id, plot in self.abc_aggregator.plot_abc_df.groupby('ID'):
            plot_sizes_un = np.unique(plot['plot_size'])
            vector_dict = {}

            if plot_sizes_un.size > 1:      # it is a nested plot, so we need to extrapolate
                nest_record_idx = plot['plot_size'] == plot_sizes_un.min(initial=5)
                small_record_idx = plot['height'] < nested_height_thresh

                for field in fields_to_summarise:   # find extrapolated summary statistics for fields_to_summarise
                    v = np.array(plot[field])
                    nest_v = v[nest_record_idx & ~small_record_idx]
                    small_v = v[nest_record_idx & small_record_idx]
                    out_v = v[~nest_record_idx]
                    sum_v = out_v.sum() + nest_v.sum() + (small_v.sum() * ((plot_sizes_un.max(initial=5.) / plot_sizes_un.min(initial=5.)) ** 2))
                    mean_v = out_v.tolist() + nest_v.tolist() + (small_v.tolist() * int((plot_sizes_un.max(initial=5.) / plot_sizes_un.min(initial=5.)) ** 2))
                    vector_dict[field] = {'sum_v': sum_v, 'mean_v': np.array(mean_v).mean(), 'n': len(mean_v)} # out_v.size + nest_v.size + (small_v.size * (plot_sizes_un.max()/ plot_sizes_un.min()) ** 2)
            else:
                for field in fields_to_summarise:
                    v = np.array(plot[field])
                    vector_dict[field] = {'sum_v': v.sum(), 'mean_v':  v.mean(), 'n': v.size}

            summary_plot = OrderedDict()
            summary_plot['ID'] = plot_id
            summary_plot['Stratum'] = plot['degr_class'].values[0]
            summary_plot['Size'] = plot_sizes_un.max(initial=5.)
            summary_plot['N'] = vector_dict['height']['n']
            summary_plot['Vol'] = vector_dict['vol']['sum_v']
            summary_plot['VolHa'] = (100. ** 2) * vector_dict['vol']['sum_v'] / (plot_sizes_un.max(initial=5.) ** 2)
            summary_plot['Height'] = vector_dict['height']['sum_v']
            summary_plot['HeightHa'] = (100. ** 2) * vector_dict['height']['sum_v'] / (plot_sizes_un.max(initial=5.) ** 2)
            summary_plot['MeanHeight'] = vector_dict['height']['mean_v']
            summary_plot['Abc'] = vector_dict['yc']['sum_v']
            summary_plot['AbcHa'] = (100. ** 2) * vector_dict['yc']['sum_v'] / (plot_sizes_un.max(initial=5.) ** 2)

            if len(self._plot_litter_df) == 0:
                logger.warning('No litter data')
            else:
                if plot_id in self._plot_litter_df.index and self._plot_litter_df.loc[plot_id, 'dry_weight'] > 0.:
                    summary_plot['LitterC'] = biomass_to_carbon_w * self._plot_litter_df.loc[plot_id, 'dry_weight'] / 1000.  # g to kg
                    # The litter quadrats are uniform across all plot sizes.  The dry weight is converted to carbon using  biomass_to_carbon_w
                    summary_plot['LitterCHa'] = summary_plot['LitterC'] * (100. ** 2) / (4 * (0.5 ** 2))
                    summary_plot['AgcHa'] = summary_plot['LitterCHa'] + summary_plot['AbcHa']
                else:
                    logger.warning('No litter data for Plot ID: {0}'.format(plot_id))
                    summary_plot['LitterC'] = np.nan
                    summary_plot['LitterCHa'] = np.nan
                    summary_plot['AgcHa'] = summary_plot['AbcHa']

            plot_summary_agc_dict[plot_id] = summary_plot

        self.plot_summary_agc_df = pd.DataFrame.from_dict(plot_summary_agc_dict, orient='index')
        return self.plot_summary_agc_df

    def write_abc_plant_file(self, out_file_name=None):
        """ Write aggregated plant ABC etc values to CSV file.

        Parameters
        ----------
        out_file_name : str
            (optional) name of csv file to write to
        """
        if len(self.abc_aggregator.plot_abc_df) == 0:
            raise Exception('There is no ABC data - call estimate() first')
        self.abc_aggregator.write_file(out_file_name=out_file_name)

    def write_agc_plot_file(self, out_file_name=None):
        """ Write AGC values etc per plot to CSV file.

        Parameters
        ----------
        out_file_name : str
            (optional) name of csv file to write to
        """

        if len(self.plot_summary_agc_df) == 0:
            raise Exception('There is no AGC data - call estimate() first')

        if out_file_name is None:
            out_file_name = str.format(str('{0}/{1} - Summary WoodyC & LitterC.csv'), os.path.dirname(self._woody_file_name),
                                       os.path.splitext(os.path.basename(self._woody_file_name))[0])

        logger.info('Writing plot AGC summary to: {0}'.format(out_file_name))
        self.plot_summary_agc_df.to_csv(out_file_name, index=False)
